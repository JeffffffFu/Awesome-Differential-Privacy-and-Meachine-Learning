from data.fed_data_distribution.pathological_nonIID_data import pathological_split_noniid
from data.get_data import get_data
from data.util.compute_model_l2_norm import compute_model_l2norm
from data.util.sampling import get_data_loaders_uniform_without_replace
from data.util.weight_initialization import init_weights
from model.CNN import CNN_tanh, CNN, CIFAR10_CNN
from model.ResNet import resnet20
from optimizer.dp_optimizer import DPSGD, DPAdam
import torch
from torch import nn
from torch.utils.data import TensorDataset
from torch.utils.data import DataLoader

from privacy_analysis.RDP.compute_rdp import compute_rdp
from privacy_analysis.RDP.rdp_convert_dp import compute_eps
from train_and_validation.train_with_dp import train_dynamic_add_noise,train_dynamic_add_noise_geo
from train_and_validation.validation import validation, validation_geo
from openpyxl import Workbook
import time
from torch_geometric.data import DataLoader as GeoDataLoader


def centralization_train_with_dp(train_data, test_data, model, batch_size, numEpoch, learning_rate, momentum, delta,
                                 max_norm, sigma, use_gnn=False):
    optimizer = DPSGD(
        l2_norm_clip=max_norm,  # 裁剪范数
        noise_multiplier=sigma,
        minibatch_size=batch_size,  # 几个样本梯度进行一次梯度下降
        microbatch_size=1,  # 几个样本梯度进行一次裁剪，这里选择逐样本裁剪
        params=model.parameters(),
        lr=learning_rate,
        momentum=momentum
    )

    # 包装抽样函数
    minibatch_size = batch_size  # 这里比较好的取值是根号n，n为每个客户端的样本数
    microbatch_size = 1  # 这里默认1就好
    iterations = 1  # n个batch，这边就定一个，每次训练采样一个Lot
    minibatch_loader, microbatch_loader = get_data_loaders_uniform_without_replace(minibatch_size, microbatch_size,
                                                                                   iterations, use_gnn=use_gnn)  # 无放回均匀采样

    # train数据在下面进行采样，这边不做dataloader
    if not use_gnn:
        test_dl = torch.utils.data.DataLoader(
            test_data, batch_size=batch_size, shuffle=False)
    else:
        test_dl = GeoDataLoader(test_data, batch_size=batch_size, shuffle=False)

    print("------ Centralized Model ------")
    rdp = 0
    orders = (list(range(2, 64)) + [128, 256, 512])  # 默认的lamda
    epsilon_list = []
    result_loss_list = []
    result_acc_list = []
    for epoch in range(numEpoch):

        train_dl = minibatch_loader(train_data)  # 抽样

        # 这里要动态加噪，每次传入的sigma可能会改变
        if not use_gnn:
            central_train_loss, central_train_accuracy = train_dynamic_add_noise(model, train_dl, optimizer)
            central_test_loss, central_test_accuracy = validation(model, test_dl)
        else:
            central_train_loss, central_train_accuracy = train_dynamic_add_noise_geo(model, train_dl, optimizer)
            central_test_accuracy = validation_geo(model, test_dl)
            print(f"epoch:{epoch}; train loss:{central_train_loss}; test acc:{central_test_accuracy:.4f}")
        # 这里要每次根据simga累加它的RDP，循环结束再转为eps，这里的epoch系数直接设为iterations(epoch里的迭代次数)，每次算一轮累和
        if not use_gnn:
            rdp_every_epoch=compute_rdp(batch_size/len(train_data), sigma, 1*iterations, orders)
        else:
            rdp_every_epoch=compute_rdp(batch_size/train_data.data.num_nodes, sigma, 1*iterations, orders)
        rdp=rdp+rdp_every_epoch
        epsilon, best_alpha = compute_eps(orders, rdp, delta)
        epsilon_list.append(epsilon)

        # result_loss_list.append(central_test_loss)
        # result_acc_list.append(central_test_accuracy)

        print("epoch: {:3.0f}".format(epoch + 1) + " | epsilon: {:7.4f}".format(
        epsilon) + " | best_alpha: {:7.4f}".format(best_alpha)  )
        #
        # print(compute_model_l2norm(model))
        # if (epsilon > 3):
        #     break
        if (epoch > 6200):
            break
        #     wb = Workbook()
        #     sheet = wb.active
        #     sheet.title = "result"
        #     sheet.cell(1, 3).value = "acc"
        #     sheet.cell(1, 4).value = "eps"
        #     sheet.cell(1, 5).value = "sigma"
        #     # sheet.cell(1, 6).value="实验时间：{}".format(datetime.datetime.now())
        #     sheet.cell(1, 7).value = "| batch_size:{}".format(batch_size) + "| learning_rate:{}".format(
        #         learning_rate) + "| sigma:{}".format(sigma) + "| max_norm:{}".format(max_norm) + "| numepoch:{}".format(
        #         numEpoch)
        #     # sheet.cell(1, 8).value = "mnist数据自适应范数裁剪，不是逐层"
        #     for i in range(len(result_loss_list)):
        #         sheet.cell(i + 2, 2).value = result_loss_list[i]
        #         sheet.cell(i + 2, 3).value = result_acc_list[i]
        #         sheet.cell(i + 2, 4).value = epsilon_list[i]
        #     wb.save("../result/{}.xlsx".format(int(time.time())))
        #     break

    print("------ Training finished ------")


if __name__ == "__main__":
    # train_data, test_data = get_data('mnist', augment=False)
    #  print(train_data.__dict__)

    from torch_geometric.datasets import Planetoid, TUDataset

    dataset = Planetoid(root='/tmp/Cora', name='Cora')

    from model.GCN import GCN

    model = GCN(dataset.num_node_features, dataset.num_classes)

    # init_weights(model, init_type='xavier', init_gain=0.5)

    # model= resnet20(10, False)
    # model= CIFAR10_CNN(3, input_norm=None, num_groups=None, size=None)
    batch_size = 2708
    learning_rate = 0.5
    numEpoch = 15000
    sigma = 1.23
    momentum = 0.9
    delta = 10 ** (-5)
    max_norm = 0.1

    # batch_size = 2708
    # learning_rate = 0.01
    # numEpoch = 15000
    # sigma = 1.23
    # momentum = 0.9
    # delta = 10 ** (-5)
    # max_norm = 0.1
    centralization_train_with_dp(dataset, dataset, model, batch_size, numEpoch, learning_rate, momentum, delta,
                                 max_norm, sigma,use_gnn=True)
