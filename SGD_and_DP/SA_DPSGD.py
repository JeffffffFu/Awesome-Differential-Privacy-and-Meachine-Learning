from data.fed_data_distribution.pathological_nonIID_data import dividing_validation_set
from data.get_data import get_data
from data.util.sampling import get_data_loaders_uniform_without_replace
from model.CNN import CNN_tanh, CNN, CIFAR10_CNN
from optimizer.dp_optimizer import DPSGD, DPAdam
import torch
from torch import nn
from torch.utils.data import TensorDataset
from torch.utils.data import DataLoader

from privacy_analysis.RDP.compute_rdp import compute_rdp
from privacy_analysis.RDP.rdp_convert_dp import compute_eps
from train_and_validation.train_with_dp import train_dynamic_add_noise
from train_and_validation.validation import validation
from openpyxl import Workbook
import time
import copy
from scipy.stats import bernoulli
import numpy as np


def centralization_train_dynamic_add_noise_SimulatedAnnealing(train_data, test_data, model,batch_size, numEpoch, learning_rate,momentum,delta,max_norm,sigma,k,mu_0):
    optimizer = DPSGD(
        l2_norm_clip=max_norm,      #裁剪范数
        noise_multiplier=sigma,
        minibatch_size=batch_size,        #几个样本梯度进行一次梯度下降
        microbatch_size=1,                #几个样本梯度进行一次裁剪，这里选择逐样本裁剪
        params=model.parameters(),
        lr=learning_rate,
        # momentum=momentum
    )

    # 包装抽样函数
    minibatch_size = batch_size  # 这里比较好的取值是根号n，n为每个客户端的样本数
    microbatch_size = 1  #这里默认1就好
    iterations = 1  # n个batch，这边就定一个，每次训练采样一个Lot
    minibatch_loader, microbatch_loader = get_data_loaders_uniform_without_replace(minibatch_size, microbatch_size, iterations)     #无放回均匀采样



    #train数据在下面进行采样，这边不做dataloader
    test_dl = torch.utils.data.DataLoader(
        test_data, batch_size=batch_size, shuffle=False)

    print("------ Centralized Model ------")
    rdp=0
    orders = (list(range(2, 64)) + [128, 256, 512])  # 默认的lamda
    epsilon_list=[]
    result_loss_list=[]
    result_acc_list=[]
    last_central_test_loss=100000.0
    last_centralized_model=model
    t=0

    for epoch in range(numEpoch):

        print("epoch:",epoch)

        train_dl = minibatch_loader(train_data)     #抽样

        central_train_loss, central_train_accuracy = train_dynamic_add_noise(model, train_dl, optimizer)

        central_test_loss, central_test_accuracy = validation(model, test_dl)


        if central_test_loss-last_central_test_loss>0:
            p = np.exp(-(central_test_loss - last_central_test_loss) * k * t)  # 概率p随着epoch增大而减小，随着损失值的差增大而减小，看出来是接收的概率
            print("central_test_loss {:7.4f}".format(central_test_loss) + "|  last_central_test_loss {:7.4f}".format(
                last_central_test_loss) + "|  接受新模型的概率p:", p)
            rv = bernoulli.rvs(p)  # 伯努利分布，p的概率输出1，1-p的概率输出0。输出1表示接受新模型

            if rv==0 and mu<mu_0 :      #==1接受模型，那么==0就是不接受模型更新，返回到上一次
                print("不接受模型更新,mu:",mu)
                #把上一轮模型保存的参数传回来
                model.load_state_dict(last_centralized_model.state_dict(),strict=True)
                mu=mu+1
            else:

                last_central_test_loss = central_test_loss
                last_central_test_acc = central_test_accuracy

                #copy.deepcopy可以只拷贝model里面的值，而不会随着model变化而改变
                last_centralized_model = copy.deepcopy(model)
                t=t+1
                print("接受新模型，当前接受的迭代轮次t：", format(t))
                mu=0

        else:
            last_central_test_loss = central_test_loss
            last_central_test_acc = central_test_accuracy
            last_centralized_model = copy.deepcopy(model)
            t = t + 1
            mu = 0
            print("接受新模型，当前接受的迭代轮次t：", format(t))


        # 这里要每次根据simga累加它的RDP，循环结束再转为eps，这里的epoch系数直接设为t,表示成功迭代的次数
        # rdp_every_epoch=compute_rdp(batch_size/len(train_data), sigma, t, orders)
        # rdp=rdp+rdp_every_epoch
        # epsilon, best_alpha = compute_eps(orders, rdp, delta)
        # epsilon_list.append(epsilon)
        #
        # result_loss_list.append(central_test_loss)
        # result_acc_list.append(central_test_accuracy)


        # print("epoch: {:3.0f}".format(epoch + 1) + " | epsilon: {:7.4f}".format(
        # epsilon) + " | best_alpha: {:7.4f}".format(best_alpha)  )

        # if (epsilon > 3):
        #     wb = Workbook()
        #     sheet = wb.active
        #     sheet.title = "result"
        #     sheet.cell(1, 3).value = "acc"
        #     sheet.cell(1, 4).value = "eps"
        #     sheet.cell(1, 5).value = "sigma"
        #     # sheet.cell(1, 6).value="实验时间：{}".format(datetime.datetime.now())
        #     sheet.cell(1, 7).value = "| batch_size:{}".format(batch_size) + "| learning_rate:{}".format(
        #         learning_rate) + "| sigma:{}".format(sigma) + "| max_norm:{}".format(max_norm) + "| numepoch:{}".format(
        #         numEpoch)
        #     # sheet.cell(1, 8).value = "mnist数据自适应范数裁剪，不是逐层"
        #     for i in range(len(result_loss_list)):
        #         sheet.cell(i + 2, 2).value = result_loss_list[i]
        #         sheet.cell(i + 2, 3).value = result_acc_list[i]
        #         sheet.cell(i + 2, 4).value = epsilon_list[i]
        #     wb.save("../result/{}.xlsx".format(int(time.time())))
        #     break

    print("------ Training finished ------")

#从训练集中划分一部分验证集来计算Loss
def centralization_train_dynamic_add_noise_SimulatedAnnealing2(train_data, test_data, model, batch_size, numEpoch,
                                                              learning_rate, momentum, delta, max_norm, sigma, k, mu_0):
    optimizer = DPSGD(
        l2_norm_clip=max_norm,  # 裁剪范数
        noise_multiplier=sigma,
        minibatch_size=batch_size,  # 几个样本梯度进行一次梯度下降
        microbatch_size=1,  # 几个样本梯度进行一次裁剪，这里选择逐样本裁剪
        params=model.parameters(),
        lr=learning_rate,
        momentum=momentum
    )

    # 包装抽样函数
    minibatch_size = batch_size  # 这里比较好的取值是根号n，n为每个客户端的样本数
    microbatch_size = 1  # 这里默认1就好
    iterations = 1  # n个batch，这边就定一个，每次训练采样一个Lot
    minibatch_loader, microbatch_loader = get_data_loaders_uniform_without_replace(minibatch_size, microbatch_size,
                                                                                   iterations)  # 无放回均匀采样

    # 从训练集样本中划分训练集和验证集，可能和通道转换有关
    # 不能在验证那边进行维度转换，那边是没有错的，现在来看通过dataload是可以自动完成维度转换的才对。
    train_data,valid_data = dividing_validation_set(train_data,4000)

    print(train_data)
    # train_data = client_data[0]
    # valid_data = client_data[1]

    # train数据在下面进行采样，这边不做dataloader
    # train_dl = torch.utils.data.DataLoader(
    #     train_data, batch_size=batch_size, shuffle=False,sampler=train_sampler)

    # train数据在下面进行采样，这边不做dataloader
    valid_dl = torch.utils.data.DataLoader(
        valid_data, batch_size=batch_size, shuffle=False)

    test_dl = torch.utils.data.DataLoader(
        test_data, batch_size=batch_size, shuffle=False)

    print("------ Centralized Model ------")
    rdp = 0
    orders = (list(range(2, 64)) + [128, 256, 512])  # 默认的lamda
    epsilon_list = []
    result_loss_list = []
    result_acc_list = []
    last_central_valid_loss = 100000.0
    last_centralized_model = model
    t = 0

    for epoch in range(numEpoch):

        print("epoch:",epoch)

        train_dl = minibatch_loader(train_data)  # 抽样

        central_train_loss, central_train_accuracy = train_dynamic_add_noise(model, train_dl, optimizer)


        central_valid_loss, central_valid_accuracy = validation(model, valid_dl) #验证集评估

        if t>11500:
            central_test_loss, central_test_accuracy = validation(model, test_dl) #测试集评估


        if central_valid_loss - last_central_valid_loss > 0:
            p = np.exp(-(central_valid_loss - last_central_valid_loss) * k * t)  # 概率p随着epoch增大而减小，随着损失值的差增大而减小，随着K的增大而减小，看出来是接收的概率
            print("central_test_loss {:7.4f}".format(central_valid_loss) + "|  last_central_test_loss {:7.4f}".format(
                last_central_valid_loss) + "|  接受新模型的概率p:", p)
            rv = bernoulli.rvs(p)  # 伯努利分布，p的概率输出1，1-p的概率输出0。输出1表示接受新模型

            if rv == 0 and mu < mu_0:  # ==1接受模型，那么==0就是不接受模型更新，返回到上一次
                print("不接受模型更新,mu:", mu)
                # 把上一轮模型保存的参数传回来
                model.load_state_dict(last_centralized_model.state_dict(), strict=True)
                mu = mu + 1
            else:

                last_central_valid_loss = central_valid_loss
                last_central_valid_acc = central_valid_accuracy

                # copy.deepcopy可以只拷贝model里面的值，而不会随着model变化而改变
                last_centralized_model = copy.deepcopy(model)
                t = t + 1
                print("接受新模型，当前接受的迭代轮次t：", format(t))
                mu = 0

        else:
            last_central_valid_loss = central_valid_loss
            last_central_valid_acc = central_valid_accuracy
            last_centralized_model = copy.deepcopy(model)
            t = t + 1
            mu = 0
            print("接受新模型，当前接受的迭代轮次t：", format(t))

        # 这里要每次根据simga累加它的RDP，循环结束再转为eps，这里的epoch系数直接设为t,表示成功迭代的次数
        # rdp_every_epoch = compute_rdp(batch_size / len(train_data), sigma, t, orders)
        # rdp = rdp + rdp_every_epoch
        # epsilon, best_alpha = compute_eps(orders, rdp, delta)
        # epsilon_list.append(epsilon)

        # result_loss_list.append(central_test_loss)
        # result_acc_list.append(central_test_accuracy)
        #
        # print("epoch: {:3.0f}".format(epoch + 1) + " | epsilon: {:7.4f}".format(
        #     epsilon) + " | best_alpha: {:7.4f}".format(best_alpha))



        if (t > 17500):
            break
        #     wb = Workbook()
        #     sheet = wb.active
        #     sheet.title = "result"
        #     sheet.cell(1, 3).value = "acc"
        #     sheet.cell(1, 4).value = "eps"
        #     sheet.cell(1, 5).value = "sigma"
        #     # sheet.cell(1, 6).value="实验时间：{}".format(datetime.datetime.now())
        #     sheet.cell(1, 7).value = "| batch_size:{}".format(batch_size) + "| learning_rate:{}".format(
        #         learning_rate) + "| sigma:{}".format(sigma) + "| max_norm:{}".format(max_norm) + "| numepoch:{}".format(
        #         numEpoch)
        #     # sheet.cell(1, 8).value = "mnist数据自适应范数裁剪，不是逐层"
        #     for i in range(len(result_loss_list)):
        #         sheet.cell(i + 2, 2).value = result_loss_list[i]
        #         sheet.cell(i + 2, 3).value = result_acc_list[i]
        #         sheet.cell(i + 2, 4).value = epsilon_list[i]
        #     wb.save("../result/{}.xlsx".format(int(time.time())))
        #     break

    print("------ Training finished ------")

if __name__=="__main__":


    #train_data, test_data = get_data('cifar10', augment=False)
    train_data, test_data = get_data('mnist', augment=False)
    #print(train_data.data)

    #model = CIFAR10_CNN(3, input_norm=None, num_groups=None, size=None)
    model = CNN_tanh()
    batch_size = 256
    learning_rate = 0.5
    numEpoch = 150000
    sigma = 1.25
    momentum = 0.9
    delta = 10 ** (-5)
    max_norm = 0.1
    k=10  #这个控制模拟退退火下接受的概率
    mu_0=5  #表示连续n次不接受新解后强制接受
    centralization_train_dynamic_add_noise_SimulatedAnnealing2(train_data, test_data, model,batch_size, numEpoch, learning_rate,momentum,delta,max_norm,sigma,k,mu_0)